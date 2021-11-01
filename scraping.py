# 必要なものをインポート
import requests
import chromedriver_binary
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import bs4
from bs4 import BeautifulSoup
import time
import pyautogui as pag
import openpyxl
import datetime

# Excelファイルを指定。
workbook = "GMBRank.xlsx" 
wb = openpyxl.load_workbook(workbook)

# Chromeをシークレットモードで用意
option = Options()                          # オプションを用意
option.add_argument('--incognito')          # シークレットモードの設定を付与
driver = webdriver.Chrome(options=option)   # Chromeを準備(optionでシークレットモードにしている）

# Googleマップを開く
url = 'https://www.google.co.jp/maps/'
driver.get(url)
driver.maximize_window()
time.sleep(5)   # 5秒待つ(以下、秒数はPCの性能に合わせる

# 店舗ごとに分かれたシートを順番に指定
for worksheet in wb.sheetnames:
    sheet = wb[worksheet]
    myshop = sheet.cell(1,1).value # myshopは店舗名
    maxcol = sheet.max_column # シートの検索ワード数を調べる
    maxrow = sheet.max_row # 書き出す行を調べる

    # Excelに日付を書き出す
    sheet.cell(maxrow + 1, 1).value = str(datetime.date.today())

    # ここで計測時間も書き出したい

    # Excelに書いてある検索ワードを順番に参照
    for keycol in range(maxcol - 1): # maxcol - 1 は検索ワード数
        keyword = sheet.cell(4, keycol + 2).value # kywoprdは検索ワード

        # 指定した検索ワードを入力し検索。
        id = driver.find_element_by_id("searchboxinput")
        id.clear()
        id.send_keys(keyword)
        time.sleep(1)

        search_button = driver.find_element_by_xpath("//*[@id='searchbox-searchbutton']")
        search_button.click()
        time.sleep(4)

        # 検索結果をスクロール。スクロールしないと8位以下が表示されない。
        # pag.moveTo(404,300)
        # pag.dragTo(404,1079,2,button="left")
        # PCによって座標の調節が必要。
        pag.moveTo(200,400)    # ドラッグからscrollに変更。
        for i in range(4):     # 座標の調節不要に。念のため4回。
            pag.scroll(-2000)
            time.sleep(0.5)


        # ここからbs4でスクレイピング
        html = driver.page_source.encode('utf-8')
        soup = BeautifulSoup(html, 'lxml')
        a = soup.find_all("a")

        # 店舗名リストを作成
        shoplist = []
        for i in range(len(a)):
            if a[i].has_attr("aria-label"):
                shoplist.append(str(a[i].attrs["aria-label"]))
        del shoplist[0:2] # 先頭2つは関係ないので削除

        # 広告は順位から除く
        if len(shoplist) <= 20: 
            for i in shoplist:
                # 一応20位までのランキングも出力しておく
                print(str(shoplist.index(i) + 1) + "." + i)
        else: # 20店舗より多い場合、先頭は広告と判定
            ad = len(shoplist) - 20
            for i in shoplist[0:ad]:
                print("広告." + i)
            for i in shoplist[ad:len(shoplist)]:
                print(str(shoplist.index(i) + 1 - ad) + "." + i)
            del shoplist[0:ad] # 広告枠をリストから削除

        # 自店の順位を調べる
        if myshop in shoplist:
            rank = str(shoplist.index(myshop) + 1)
            print(myshop + " は " + "「" + keyword + "」" + " で " + rank + "位です。" + "\n")
            # Excelに順位を書き込む
            sheet.cell(maxrow + 1, keycol + 2).value = int(rank)
        else:
            rank = "圏外"
            print(myshop + " は " + "「" + keyword + "」" + " で " + rank + "です。" + "\n")
            # Excelに順位を書き込む
            sheet.cell(maxrow + 1, keycol + 2).value = str(rank)

        # Excelを保存する
        wb.save(workbook)